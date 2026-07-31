"""Insert wholesale CAD prices into catalogue HTML."""
import openpyxl
import re
from difflib import SequenceMatcher

# ── 1. Load Excel prices ──
wb = openpyxl.load_workbook(r'D:\sahi_temp_prices.xlsx', data_only=True)
ws = wb.active
prices_by_title = {}
prices_by_sku = {}
for row in range(2, ws.max_row + 1):
    title = str(ws.cell(row=row, column=3).value or '').strip()
    price_f = ws.cell(row=row, column=6).value
    prod_id = str(ws.cell(row=row, column=2).value or '').strip()
    if title and price_f:
        try:
            retail = float(price_f)
            wholesale = round(retail / 3, 2)
            prices_by_title[title.lower().strip()] = {'retail': retail, 'wholesale': wholesale}
        except:
            pass

print(f"Excel products loaded: {len(prices_by_title)}")

# Also load with SKU in title matching
# Some Excel titles contain SKU codes like "BGZBM210-SLV", "BGZBM212-SLV"
for key in list(prices_by_title.keys()):
    # Extract potential SKU codes from title
    sku_matches = re.findall(r'\b(BG[A-Z]+[\d-]+|AP[A-Z]+[\d-]+|JW[A-Z]+[\d-]+)\b', key.upper())
    for m in sku_matches:
        prices_by_sku[m] = prices_by_title[key]

print(f"SKU-indexed entries: {len(prices_by_sku)}")

# ── 2. Read catalogue HTML ──
with open(r'D:\sahi_recovery\public\catalogue.html', 'r', encoding='utf-8') as f:
    html = f.read()

# ── 3. Normalize title for matching ──
def normalize(t):
    t = t.lower().strip()
    t = re.sub(r'\bsahi\s+london\b', '', t)
    t = re.sub(r'\bpop art\b', '', t)
    t = re.sub(r'\bwomens\b', '', t)
    t = re.sub(r'\bcollection\b', '', t)
    t = re.sub(r'\blimited edition\b', '', t)
    t = ' '.join(t.split())
    return t

def find_price(search_text, tile_html):
    """Find wholesale price by matching tile text with Excel data."""
    # Try direct title match
    key = search_text.lower().strip()
    if key in prices_by_title:
        return prices_by_title[key]['wholesale']

    # Try normalized fuzzy match
    cat_norm = normalize(search_text)
    
    best_score = 0
    best_price = None
    for ek, ed in prices_by_title.items():
        ek_norm = normalize(ek)
        if cat_norm in ek_norm or ek_norm in cat_norm:
            return ed['wholesale']
        score = SequenceMatcher(None, cat_norm, ek_norm).ratio()
        if score > best_score:
            best_score = score
            best_price = ed['wholesale']

    if best_score > 0.55:
        return best_price

    # Try SKU-based matching (for bags: extract base model from variant SKU)
    sku_match = re.search(r'<div class="tile-sku">(.*?)</div>', tile_html, re.DOTALL)
    if sku_match:
        tile_sku = sku_match.group(1).strip()
        # Try full SKU first
        if tile_sku in prices_by_sku:
            return prices_by_sku[tile_sku]['wholesale']
        # Try base SKU (remove variant: BGUAWRF201-1 → BGUAWRF201)
        base = re.sub(r'-\d+$', '', tile_sku)
        if base != tile_sku and base in prices_by_sku:
            return prices_by_sku[base]['wholesale']
    
    # Try extracting base SKU from variant SKUs
    variant_matches = re.findall(r'<div class="variant-sku">(.*?)</div>', tile_html)
    if variant_matches:
        for v in variant_matches:
            base = v.split('-')[0] if '-' in v else v
            if base in prices_by_sku:
                return prices_by_sku[base]['wholesale']

    return None

# ── 4. Process each tile ──
modified_html = html
tile_re = re.compile(r'(<div class="tile">.*?</div>\s*</div>)\s*(?=<div class="tile">|</div>\s*</section>|$)', re.DOTALL)

tiles = tile_re.findall(html)
matched = 0
unmatched = []

for tile_html in tiles:
    # Extract search text
    name_match = re.search(r'<div class="tile-name">(.*?)</div>', tile_html, re.DOTALL)
    sku_match = re.search(r'<div class="tile-sku">(.*?)</div>', tile_html, re.DOTALL)
    alt_match = re.search(r'alt="(.*?)"', tile_html)
    
    search_text = None
    if name_match:
        search_text = name_match.group(1).strip()
    elif sku_match:
        search_text = sku_match.group(1).strip()
    elif alt_match:
        search_text = alt_match.group(1).strip()
    
    if not search_text:
        unmatched.append('NO TEXT')
        continue
    
    price = find_price(search_text, tile_html)
    
    if price:
        # Add price line to tile - after tile-name or tile-sku
        price_html = f'<div class="tile-price">Wholesale CAD ${price:.2f}</div>'
        
        if name_match:
            # Insert after tile-name
            old_name_line = name_match.group(0)
            new_block = old_name_line + '\n      ' + price_html
            modified_html = modified_html.replace(old_name_line, new_block, 1)
        elif sku_match:
            # Insert after tile-sku
            old_sku_line = sku_match.group(0)
            new_block = old_sku_line + '\n      ' + price_html
            modified_html = modified_html.replace(old_sku_line, new_block, 1)
        
        matched += 1
        print(f'  CAD ${price:.2f} → {search_text[:55]}')
    else:
        unmatched.append(search_text[:55])
        print(f'  [NO MATCH] {search_text[:55]}')

print(f'\nMatched: {matched}/{len(tiles)}')
print(f'Unmatched: {len(unmatched)}')

# ── 5. Add CSS for tile-price ──
price_css = '\n  .tile-price {\n    font-size: 12px;\n    font-weight: 600;\n    color: #1A1A1A;\n    letter-spacing: 0.5px;\n    margin-top: 6px;\n    padding: 4px 10px;\n    display: inline-block;\n    background: #F5F0EB;\n    border-radius: 2px;\n  }'

# Insert CSS before the first </style> or after specific location
if '.tile-sku {' in modified_html:
    modified_html = modified_html.replace('.tile-sku {', price_css + '\n  .tile-sku {')

# ── 6. Write output ──
with open(r'D:\sahi_recovery\public\catalogue.html', 'w', encoding='utf-8') as f:
    f.write(modified_html)

print('\nCatalogue updated with wholesale prices!')
